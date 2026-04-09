"""Export service for generating participant exports in Excel/CSV/PDF formats."""

import csv
import io
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from event.models import CustomField, Event, Participant


# Column definitions: (id, german_label, type, value_extractor)
STANDARD_COLUMNS = [
    ("first_name", "Vorname", "standard"),
    ("last_name", "Nachname", "standard"),
    ("scout_name", "Pfadfindername", "standard"),
    ("email", "E-Mail", "standard"),
    ("birthday", "Geburtstag", "standard"),
    ("age", "Alter", "computed"),
    ("gender", "Geschlecht", "standard"),
    ("address", "Adresse", "standard"),
    ("zip_code", "PLZ", "standard"),
    ("city", "Stadt", "standard"),
    ("booking_option", "Buchungsoption", "standard"),
    ("is_paid", "Bezahlt", "computed"),
    ("total_paid", "Bezahlter Betrag", "computed"),
    ("remaining_amount", "Offener Betrag", "computed"),
    ("payment_method", "Zahlungsmethode", "computed"),
    ("nutritional_tags", "Ernährungstags", "computed"),
    ("labels", "Labels", "computed"),
]


def _compute_age(birthday, reference_date: date) -> str:
    """Calculate age at reference date."""
    if not birthday:
        return ""
    age = reference_date.year - birthday.year
    if (reference_date.month, reference_date.day) < (birthday.month, birthday.day):
        age -= 1
    return str(age)


def _get_column_value(participant: Participant, col_id: str, reference_date: date) -> str:
    """Extract the value for a given column from a participant."""
    if col_id == "first_name":
        return participant.first_name
    if col_id == "last_name":
        return participant.last_name
    if col_id == "scout_name":
        return participant.scout_name
    if col_id == "email":
        return participant.email
    if col_id == "birthday":
        return participant.birthday.strftime("%d.%m.%Y") if participant.birthday else ""
    if col_id == "age":
        return _compute_age(participant.birthday, reference_date)
    if col_id == "gender":
        return participant.get_gender_display()
    if col_id == "address":
        return participant.address
    if col_id == "zip_code":
        return participant.zip_code
    if col_id == "city":
        return participant.city
    if col_id == "booking_option":
        return participant.booking_option.name if participant.booking_option else ""
    if col_id == "is_paid":
        return "Ja" if participant.is_paid else "Nein"
    if col_id == "total_paid":
        return f"{participant.total_paid:.2f}"
    if col_id == "remaining_amount":
        return f"{participant.remaining_amount:.2f}"
    if col_id == "payment_method":
        last_payment = participant.payments.order_by("-received_at").first()
        if last_payment:
            return str(last_payment.get_method_display())
        return ""
    if col_id == "nutritional_tags":
        tags = participant.nutritional_tags.all()
        return ", ".join(t.name for t in tags)
    if col_id == "labels":
        return ", ".join(l.name for l in participant.labels.all())

    # Custom field value
    if col_id.startswith("cf_"):
        cf_id = int(col_id[3:])
        from event.models import CustomFieldValue

        try:
            cfv = CustomFieldValue.objects.get(custom_field_id=cf_id, participant=participant)
            return cfv.value
        except CustomFieldValue.DoesNotExist:
            return ""

    return ""


class ExportService:
    """Service for exporting event participants."""

    @staticmethod
    def get_available_columns(event: Event) -> list[dict]:
        """Return all available columns for an event, including custom fields."""
        columns = []
        for col_id, label, col_type in STANDARD_COLUMNS:
            columns.append(
                {
                    "id": col_id,
                    "label": label,
                    "type": col_type,
                }
            )

        # Add custom field columns
        for cf in CustomField.objects.filter(event=event).order_by("sort_order"):
            columns.append(
                {
                    "id": f"cf_{cf.id}",
                    "label": cf.label,
                    "type": "custom_field",
                }
            )

        return columns

    @staticmethod
    def _get_participants(event: Event, filters: dict | None = None) -> list[Participant]:
        """Get filtered participants for export."""
        qs = (
            Participant.objects.filter(registration__event=event)
            .select_related("booking_option", "registration")
            .prefetch_related("nutritional_tags", "labels", "payments")
            .order_by("last_name", "first_name")
        )

        if filters:
            if filters.get("is_paid") is True:
                # Filter for paid — done in Python since is_paid is computed
                pass
            if filters.get("is_paid") is False:
                pass
            if filters.get("booking_option_id"):
                qs = qs.filter(booking_option_id=filters["booking_option_id"])
            if filters.get("label_id"):
                qs = qs.filter(labels__id=filters["label_id"]).distinct()

        participants = list(qs)

        # Apply computed filters in Python
        if filters and filters.get("is_paid") is True:
            participants = [p for p in participants if p.is_paid]
        elif filters and filters.get("is_paid") is False:
            participants = [p for p in participants if not p.is_paid]

        return participants

    @staticmethod
    def _resolve_columns(columns: list[str], event: Event) -> list[tuple[str, str]]:
        """Resolve column IDs to (id, label) pairs. 'all' expands to all columns."""
        available = ExportService.get_available_columns(event)
        available_map = {c["id"]: c["label"] for c in available}

        if "all" in columns:
            return [(c["id"], c["label"]) for c in available]

        result = []
        for col_id in columns:
            label = available_map.get(col_id, col_id)
            result.append((col_id, label))
        return result

    @staticmethod
    def export_participants(
        event: Event,
        columns: list[str],
        fmt: str,
        filters: dict | None = None,
    ) -> tuple[bytes, str, str]:
        """
        Export participants to the requested format.

        Returns: (file_bytes, content_type, filename)
        """
        resolved_columns = ExportService._resolve_columns(columns, event)
        participants = ExportService._get_participants(event, filters)

        reference_date = event.start_date or date.today()

        if fmt == "excel":
            return ExportService._export_excel(event, resolved_columns, participants, reference_date)
        if fmt == "csv":
            return ExportService._export_csv(event, resolved_columns, participants, reference_date)
        if fmt == "pdf":
            return ExportService._export_pdf(event, resolved_columns, participants, reference_date)

        raise ValueError(f"Unbekanntes Exportformat: {fmt}")

    @staticmethod
    def _export_excel(
        event: Event,
        columns: list[tuple[str, str]],
        participants: list[Participant],
        reference_date: date,
    ) -> tuple[bytes, str, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Teilnehmer"

        # Header row
        for col_idx, (col_id, label) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            from copy import copy

            font = copy(cell.font)
            font.bold = True
            cell.font = font

        # Data rows
        for row_idx, participant in enumerate(participants, 2):
            for col_idx, (col_id, _) in enumerate(columns, 1):
                value = _get_column_value(participant, col_id, reference_date)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col_idx, (col_id, label) in enumerate(columns, 1):
            max_len = len(label)
            for row_idx in range(2, len(participants) + 2):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        slug = event.slug
        return (
            buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"teilnehmer-{slug}.xlsx",
        )

    @staticmethod
    def _export_csv(
        event: Event,
        columns: list[tuple[str, str]],
        participants: list[Participant],
        reference_date: date,
    ) -> tuple[bytes, str, str]:
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_ALL)

        # Header
        writer.writerow([label for _, label in columns])

        # Data
        for participant in participants:
            row = [_get_column_value(participant, col_id, reference_date) for col_id, _ in columns]
            writer.writerow(row)

        # UTF-8 BOM for Excel compatibility
        content = buf.getvalue()
        bom = "\ufeff"
        file_bytes = (bom + content).encode("utf-8")

        slug = event.slug
        return (
            file_bytes,
            "text/csv; charset=utf-8",
            f"teilnehmer-{slug}.csv",
        )

    @staticmethod
    def _export_pdf(
        event: Event,
        columns: list[tuple[str, str]],
        participants: list[Participant],
        reference_date: date,
    ) -> tuple[bytes, str, str]:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(f"Teilnehmerliste: {event.name}", styles["Title"]))
        elements.append(Spacer(1, 0.5 * cm))

        # Table data: checkbox column + selected columns
        header = ["\u2610"] + [label for _, label in columns]
        table_data = [header]

        for participant in participants:
            row = ["\u2610"]  # Checkbox
            for col_id, _ in columns:
                value = _get_column_value(participant, col_id, reference_date)
                row.append(value)
            table_data.append(row)

        if len(table_data) == 1:
            elements.append(Paragraph("Keine Teilnehmer gefunden.", styles["Normal"]))
        else:
            # Calculate column widths
            available_width = landscape(A4)[0] - 3 * cm
            n_cols = len(header)
            checkbox_width = 1 * cm
            data_col_width = (available_width - checkbox_width) / max(n_cols - 1, 1)
            col_widths = [checkbox_width] + [data_col_width] * (n_cols - 1)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8e4f0")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#4a3a6b")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f6ff")]),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            elements.append(table)

        doc.build(elements)
        buf.seek(0)

        slug = event.slug
        return (
            buf.getvalue(),
            "application/pdf",
            f"teilnehmer-{slug}.pdf",
        )
